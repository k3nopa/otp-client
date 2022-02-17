import openpyxl

excel_file = "exp_results.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb['a_1000_n_128_h_8']
target = 60

result_file = "results.txt"
with open(result_file, "r") as f:
    lines = f.readlines()
    target_cell = 65    # 'A'
    current_row = 1
    sample_counter = 0

    for line in lines:
        if ':' not in line: # Skip final avg in table.
            data = line.strip().split('\t')
            if len(data) > 1:   # To avoid array with '' elements.
                ws[f"{chr(target_cell)}{current_row}"] = data[0]
                if data[-1].isnumeric():
                    ws[f"{chr(target_cell+1)}{current_row}"] = int(data[-1])
                else: 
                    if '.' in data[-1]:
                        ws[f"{chr(target_cell+1)}{current_row}"] = float(data[-1])
                    else:
                        ws[f"{chr(target_cell+1)}{current_row}"] = data[-1]
                if sample_counter == target:
                    current_row  = 1
                    sample_counter = 0
                else:
                    current_row += 1
                    sample_counter += 1
        else:
            target_cell += 2

wb.save(excel_file)
